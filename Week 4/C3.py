# Title:Reading data from .xlsx file (Excel File) in to Tree view.
# Description: Python GUI Programming: This code extracts the datate from an excel file and uses Tree view to show the files.
# Copyright: Rajvir Singh
# Author: Rajvir Singh
# 13-08-2021
# Version: 1.00	     

import xlrd 
import openpyxl
import tkinter as tk
from tkinter import ttk
from tkinter.messagebox import showinfo

xlrd.xlsx.ensure_elementtree_imported(False, None)
xlrd.xlsx.Element_has_iter = True
 
#The GUI Window
root = tk.Tk()
root.title('Treeview demo')
root.geometry('620x200')

#The file to be opened 
loc = (r"excel.xlsx")
  
# To open Workbook 
wb = xlrd.open_workbook(loc)
worksheet = wb.sheet_by_name('Sheet1') 
worksheet = wb.sheet_by_index(0) 

# To read and print the file contents
i=1
data=[]
while (i<11):
 sn = worksheet.cell(i, 0)
 rn = worksheet.cell(i, 1)
 mrks = worksheet.cell(i, 2)
 print(" ", str(sn.value), " , ", str(rn.value), " , ",str(mrks.value))
 data.append((str(sn.value), str(rn.value),str(mrks.value))) 
 i=i+1


# columns
columns = ('#1', '#2', '#3')

#Creating tree
tree = ttk.Treeview(root, columns=columns, show='headings')

# define headings
tree.heading('#1', text='SN')
tree.heading('#2', text='Student ID')
tree.heading('#3', text='Marks')



# adding data to the treeview

for d in data:
	tree.insert('', tk.END, values=d)

def item_selected(event):
    for selected_item in tree.selection():
        # dictionary
        item = tree.item(selected_item)
        # list
        record = item['values']
        #
        showinfo(title='Information',
                message=','.join(record))


tree.bind('<<TreeviewSelect>>', item_selected)

tree.grid(row=0, column=0, sticky='nsew')

# add a scrollbar
scrollbar = ttk.Scrollbar(root, orient=tk.VERTICAL, command=tree.yview)
tree.configure(yscroll=scrollbar.set)
scrollbar.grid(row=0, column=1, sticky='ns')

# run the app
root.mainloop()
